# processor.py
# Lógica de negocio: operaciones sobre Excel

from openpyxl import load_workbook


def process_excel(path):
    #Acceso a la hoja de datos de excel
    wb= load_workbook(path)
    #ws = wb["Datos"]
    #Recorrer todas las filas desde la fila 2 hasta la ultima fila con datos
    #for row in range(2,ws.max_row+1):
        # columna D: identificador limpio
        #ws[f"D{row}"] = clean_id(ws[f"A{row}"].value)
        # columna E: nombre completo
        #ws[f"E{row}"] = merge_name(ws[f"B{row}"].value,ws[f"C{row}"].value)
    # guarde los cambios en el mismo archivo
    wb.save(path)
    
def ejecutar_accion(instruccion,path):
    # Abre el archivo de ejemplo
    wb = load_workbook(path)
    ws = wb.active

    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"{col}{fila}"] = ''.join(filter(str.isdigit, str(ws[f"{col}{fila}"].value)))

    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["A" + str(fila)].value or ""
            apellido = ws["B" + str(fila)].value or ""
            ws["C" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save("ejemplo.xlsx")
