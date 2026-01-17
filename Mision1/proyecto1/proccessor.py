import re
from openpyxl import load_workbook
# =====================
# FUNCION clean_id
# Elimina caracteres no numéricos de un documento
# ""cc75.888.56" = "7588856"
# =====================
def clean_id(value):
    # elimina caracteres no numéricos de un documento
    if value is None:
        return ""
    return re.sub(r"\D", "",str(value))
# =====================
# Funcion merge_name
# Une un nombre y apellido en un solo campo
# =====================
def merge_name(name, lastname):
    # une un nombre y apellido en un solo campo y si tiene un espacio extra lo elimina con strip
    if name is None:
        name = ""
    if lastname is None:
        lastname = ""
    return f"{name} {lastname}".strip()
def process_excel(path):
    #Acceso a la hoja de datos de excel
    wb= load_workbook(path)
    ws = wb["Datos"]
    #Recorrer todas las filas desde la fila 2 hasta la ultima fila con datos
    for row in range(2,ws.max_row+1):
        # columna D: identificador limpio
        ws[f"D{row}"] = clean_id(ws[f"A{row}"].value)
        # columna E: nombre completo
        ws[f"E{row}"] = merge_name(ws[f"B{row}"].value,ws[f"C{row}"].value)
    # guarde los cambios en el mismo archivo
    wb.save(path)

def process_excel_safe(path):
    try:
        process_excel(path)
        return True, "Archivo procesado correctamente"
    except PermissionError:
        return (
            False,
            "El archivo excel esta abierto. \n"
            "Por favor cierre el archivo e intente de nuevo."
        )
    except KeyError:
        return False, "Hoja 'Datos' no encontrada"
    except Exception as e:
        return False,f"Error inesperado: {str(e)}"
        
    
    
                                